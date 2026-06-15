import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models import Miembro, Inventario


def estilo_encabezado(ws, headers):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def exportar_miembros_excel(miembros: list[Miembro]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Miembros"

    headers = [
        "Cédula", "Nombres", "Apellidos", "Fecha_Nacimiento", "Teléfono",
        "Correo_Electrónico", "Sexo", "Estado_Civil", "Estado",
        "Fecha_Bautismo", "Fecha_Conversión", "Iglesia", "Familia",
        "Ciudad", "Parroquia", "Dirección", "Profesiones", "Oficios",
    ]
    estilo_encabezado(ws, headers)

    for i, m in enumerate(miembros, 2):
        ws.cell(row=i, column=1, value=m.Cedula)
        ws.cell(row=i, column=2, value=m.Nombres)
        ws.cell(row=i, column=3, value=m.Apellidos)
        ws.cell(row=i, column=4, value=str(m.Fecha_Nacimiento) if m.Fecha_Nacimiento else "")
        ws.cell(row=i, column=5, value=m.Telefono or "")
        ws.cell(row=i, column=6, value=m.Correo_Electronico or "")
        ws.cell(row=i, column=7, value=m.Sexo or "")
        ws.cell(row=i, column=8, value=m.Estado_Civil or "")
        ws.cell(row=i, column=9, value=m.Estado)
        ws.cell(row=i, column=10, value=str(m.Fecha_Bautismo) if m.Fecha_Bautismo else "")
        ws.cell(row=i, column=11, value=str(m.Fecha_Conversion) if m.Fecha_Conversion else "")
        ws.cell(row=i, column=12, value=m.iglesia.Nombre_Iglesia if m.iglesia else "")
        ws.cell(row=i, column=13, value=m.familia.Nombre_Familia if m.familia else "")
        ws.cell(row=i, column=14, value=m.ciudad.Nombre_Ciudad if m.ciudad else "")
        ws.cell(row=i, column=15, value=m.parroquia.Nombre_Parroquia if m.parroquia else "")
        ws.cell(row=i, column=16, value=m.Direccion or "")
        ws.cell(row=i, column=17, value=", ".join(p.Nombre_Profesion for p in m.profesiones))
        ws.cell(row=i, column=18, value=", ".join(o.Nombre_Oficio for o in m.oficios))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_inventario_excel(articulos: list[Inventario]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "Artículo", "Descripción", "Marca", "Modelo", "N° Serie",
        "Cantidad", "Categoría", "Iglesia", "Estado",
        "Ubicación_Interna", "Responsable", "Tipo_Ubicación", "Ubicación_Detallada",
    ]
    estilo_encabezado(ws, headers)

    for i, a in enumerate(articulos, 2):
        ws.cell(row=i, column=1, value=a.Nombre_Articulo)
        ws.cell(row=i, column=2, value=a.Descripcion or "")
        ws.cell(row=i, column=3, value=a.Marca or "")
        ws.cell(row=i, column=4, value=a.Modelo or "")
        ws.cell(row=i, column=5, value=a.Numero_Serie or "")
        ws.cell(row=i, column=6, value=a.Cantidad)
        ws.cell(row=i, column=7, value=a.categoria.Nombre_Categoria if a.categoria else "")
        ws.cell(row=i, column=8, value=a.iglesia.Nombre_Iglesia if a.iglesia else "")
        ws.cell(row=i, column=9, value=a.Estado_Articulo)
        ws.cell(row=i, column=10, value=a.Ubicacion_Interna or "")
        ws.cell(row=i, column=11, value=f"{a.miembro_resguarda.Nombres} {a.miembro_resguarda.Apellidos}" if a.miembro_resguarda else "")
        ws.cell(row=i, column=12, value=a.Tipo_Ubicacion)
        ws.cell(row=i, column=13, value=a.Ubicacion_Detallada or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
